import pandas as pd, json
import os
import httpx
import traceback
from dotenv import load_dotenv
from openai import AzureOpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path

load_dotenv()



# MAKE CONNECTION FOR TLS handshake
timeout = httpx.Timeout(connect=45.0, read=30.0,write=30.0, pool=30.0)

http_client = httpx.Client(
    timeout=timeout,
    limits=httpx.Limits(max_connections=10, max_keepalive_connections=5),
    trust_env=True,   # safe even if you think no proxy
)

aoai_client = AzureOpenAI(
    azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT"),
    api_key = os.getenv("AZURE_OPENAI_KEY") ,
    api_version="2024-02-01",
    http_client=http_client,
    timeout=30.0,
    max_retries=1
)



def read_excel(file_path):
    df = pd.read_excel(file_path, engine="openpyxl", dtype=str)

    # Normalize column names
    df.columns = [str(col).strip() for col in df.columns]

    # Replace NaN with empty string
    df = df.fillna("")

    # Parse Date/Time safely
    if "Date/Time" not in df.columns:
        raise ValueError("Column 'Date/Time' not found in Excel file")

    df["DateParsed"] = pd.to_datetime(
        df["Date/Time"],
        format="%d.%m.%Y %H:%M:%S",
        errors="coerce"
    )

    # Fallback parsing if some rows fail
    if df["DateParsed"].isna().any():
        fallback_mask = df["DateParsed"].isna()
        df.loc[fallback_mask, "DateParsed"] = pd.to_datetime(
            df.loc[fallback_mask, "Date/Time"],
            errors="coerce",
            dayfirst=True
        )

    # Sort by parsed date
    df = df.sort_values("DateParsed", na_position="last")

    lines = ["Transactions:","-"*85,"Date/Time | Transaction | Transaction Description | Program | Table Name | Activity","-"*85]
    
    for _, r in df.iterrows():
        time_str = (
            r["DateParsed"].strftime("%H:%M:%S")
            if pd.notna(r["DateParsed"])
            else str(r.get("Date/Time", "")).strip()
        )

        transaction = str(r.get("Transaction", "")).strip()
        program = str(r.get("Program", "")).strip()
        table_name = str(r.get("Table Name", "")).strip()
        transaction_desc = str(r.get("Transaction Description", "")).strip()


        line = f"{time_str} | {transaction} | {transaction_desc} | {program} | {table_name} | "

        # Add optional useful fields only if they have data
        extra_parts = []

        activity_desc = str(r.get("Activity Description", "")).strip()
        if activity_desc:
            extra_parts.append(activity_desc)

        old_value = str(r.get("Old Value", "")).strip()
        new_value = str(r.get("New Value", "")).strip()
        if old_value or new_value:
            extra_parts.append(f"Old Value: {old_value} -> New Value: {new_value}")

        if extra_parts:
            line += " ; ".join(extra_parts)

        lines.append(line)

    return "\n".join(lines)

def call_model(prompt):
    try:
        return aoai_client.chat.completions.create(
            model= os.getenv("AZURE_OPENAI_DEPLOYMENT"),
                response_format={"type": "json_object"},  # 👈 Force valid JSON
                messages=[
                    {
                        "role": "system",
                        "content": "You are an SAP GRC Firefighter ID (FFID) log reviewer"
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                max_tokens=1200,
                temperature=0.2,
                top_p = 1,
                frequency_penalty = 0,
                presence_penalty = 0,
        )
    except:
        print(f"132 {traceback.print_exc()}")
        return {}

def append_dict_as_new_columns_first_row(file_path, data, sheet_name="Data"):
    # ---------- Styling ----------
    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")   # Dark blue
    data_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")     # Light yellow

    header_font = Font(color="FFFFFF", bold=True)
    data_font = Font(color="000000", bold=False)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # ---------- Load or create workbook ----------
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # ---------- Read existing headers ----------
    headers = {}
    max_col = ws.max_column

    # If sheet is completely empty
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        max_col = 0

    for col in range(1, max_col + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value:
            headers[header_value] = col

    # ---------- Append only missing columns ----------
    new_columns_added = []
    for key in data.keys():
        if key not in headers:
            max_col += 1
            headers[key] = max_col
            new_columns_added.append(key)

            # Write new header
            header_cell = ws.cell(row=1, column=max_col, value=key)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = center_align
            header_cell.border = thin_border

    # ---------- Write values only in row 2 for appended/new columns ----------
    target_row = 2

    for key in new_columns_added:
        col_index = headers[key]
        value = data.get(key, "")

        cell = ws.cell(row=target_row, column=col_index, value=value)
        cell.fill = data_fill
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border

    # ---------- Auto-adjust only new columns ----------
    for key in new_columns_added:
        col_index = headers[key]
        header_text = str(key)
        value_text = str(data.get(key, ""))

        max_length = max(len(header_text), len(value_text))
        adjusted_width = min(max_length + 4, 50)  # keep width reasonable

        ws.column_dimensions[ws.cell(row=1, column=col_index).column_letter].width = adjusted_width

    # ---------- Slightly increase row height for better look ----------
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[target_row].height = 35

    # ---------- Save workbook ----------
    wb.save(file_path)
    print(f"New columns appended and values written in row {target_row} without affecting existing data.")

def calculate_openai_cost_inr(
        input_tokens,
        output_tokens,
        usd_to_inr_rate=90.0,
        input_price_per_million=0.40,
        output_price_per_million=1.60
    ):
        """
        Calculate OpenAI GPT-4.1-mini API cost in INR

        :param input_tokens: Number of input (prompt) tokens
        :param output_tokens: Number of output (completion) tokens
        :param usd_to_inr_rate: USD to INR conversion rate
        :param input_price_per_million: Cost per 1M input tokens (USD)
        :param output_price_per_million: Cost per 1M output tokens (USD)
        :return: Total cost in INR
        """

        input_cost_usd = (input_tokens / 1_000_000) * input_price_per_million
        output_cost_usd = (output_tokens / 1_000_000) * output_price_per_million

        total_cost_usd = input_cost_usd + output_cost_usd
        total_cost_inr = total_cost_usd * usd_to_inr_rate

        return round(total_cost_inr, 4)

def call_llm():

    # Set your log directory path here
    log_dir = Path("Logs")   # Example: Path(r"C:\Users\YourName\log")

    # Excel file extensions to look for
    excel_extensions = {".xlsx", ".xls", ".xlsm", ".xlsb"}

    # Get full paths of all Excel files recursively
    excel_files = [
        str(file.resolve())
        for file in log_dir.rglob("*")
        if file.is_file() and file.suffix.lower() in excel_extensions
    ]

    # Print the list
    for file_path in excel_files:

    # file_path = r"C:\Users\111439\OneDrive - Torrent Gas Ltd\Desktop\selenium\Logs\export_20260409_143539.xlsx"

        transactions = read_excel(file_path)

        with open("prompt.txt","r") as file:
            prompt = file.read()

        result = call_model(prompt+transactions)

        llm_response = result.choices[0].message.content

        parsed_data = json.loads(llm_response)

        return parsed_data
        # # # ✅ Example usage (your case)
        # input_tokens = result.usage.prompt_tokens
        # output_tokens = result.usage.completion_tokens

        # cost = 0
        # cost_inr = calculate_openai_cost_inr(input_tokens, output_tokens)
        # cost = cost + cost_inr

# with open("count.txt","w") as file:
#     file.write(str(cost))

from glob import glob

def get_latest_excel_file(directory_name):
    """
    Returns the full path of the latest (most recently modified)
    Excel file in the given directory.
    """
    excel_files = glob(os.path.join(os.getcwd(),directory_name, "*.xlsx")) + \
                  glob(os.path.join(os.getcwd(),directory_name, "*.xls"))

    if not excel_files:
        return None

    return max(excel_files, key=os.path.getmtime)

excel_file = get_latest_excel_file(os.getenv('DOWNLOAD_DIR'))

transactions = read_excel(excel_file)

with open("prompt.txt","r") as file:
    prompt = file.read()

result = call_model(prompt+transactions)

llm_response = result.choices[0].message.content

parsed_data = json.loads(llm_response)

print(excel_file)
print(parsed_data)


if os.path.exists(excel_file):
    os.remove(excel_file)
    print("File deleted successfully")
else:
    print("File not found")